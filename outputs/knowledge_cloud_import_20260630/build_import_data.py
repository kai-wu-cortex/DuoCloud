from __future__ import annotations

import hashlib
import json
import re
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SPEC_XLSX = Path("/Users/kyle/Documents/Codex/2026-06-18/documents-plugin-documents-openai-primary-runtime/outputs/knowledge_cloud_spec/品特烫金膜知识云搭建Spec规范表.xlsx")
VAULT_ROOT = Path("/Users/kyle/Library/Mobile Documents/iCloud~md~obsidian/Documents/HotFoil_Database")
OUT_DIR = Path("/Users/kyle/codex project/Duo Cloud/outputs/knowledge_cloud_import_20260630")

EXCLUDE_PARTS = {".obsidian", ".git", ".claude", ".claudian", "90_模板"}

TABLE_CODE_BY_NAME = {
    "产品主数据表": "T01_PRODUCT",
    "底材知识表": "T02_SUBSTRATE",
    "产品×底材适配规则表": "T03_COMPATIBILITY",
    "产品 × 底材适配规则表": "T03_COMPATIBILITY",
    "工艺知识表": "T04_PROCESS",
    "报价规则表": "T05_QUOTE",
    "质量问题与解决方案表": "T06_DEFECT",
    "供应链能力表": "T07_SUPPLY",
    "销售话术与FAQ表": "T08_FAQ",
    "销售话术与 FAQ 表": "T08_FAQ",
    "知识标签体系表": "T09_TAG",
    "知识治理表": "T10_GOVERNANCE",
}

SHEET_TITLE_BY_CODE = {
    "T01_PRODUCT": "T01_产品主数据",
    "T02_SUBSTRATE": "T02_底材知识",
    "T03_COMPATIBILITY": "T03_适配规则",
    "T04_PROCESS": "T04_工艺知识",
    "T05_QUOTE": "T05_报价规则",
    "T06_DEFECT": "T06_质量问题",
    "T07_SUPPLY": "T07_供应链能力",
    "T08_FAQ": "T08_FAQ话术",
    "T09_TAG": "T09_标签体系",
    "T10_GOVERNANCE": "T10_知识治理",
}

META_COLUMNS = ["来源Obsidian路径", "原始标题", "来源类型", "标签", "更新时间", "入库备注"]


def s(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, list):
        return "、".join(s(v) for v in value if s(v))
    return str(value).strip().strip('"').strip("'")


def clean_text(value: str, limit: int = 1200) -> str:
    value = re.sub(r"!\[\[([^\]]+)\]\]", r"附件：\1", value)
    value = re.sub(r"\[\[([^\]|]+)\|([^\]]+)\]\]", r"\2", value)
    value = re.sub(r"\[\[([^\]]+)\]\]", r"\1", value)
    value = re.sub(r"```[\s\S]*?```", " ", value)
    value = re.sub(r"^#{1,6}\s*", "", value, flags=re.M)
    value = re.sub(r"^\s*[-*]\s*", "", value, flags=re.M)
    value = re.sub(r"\|", " / ", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value[:limit]


def stable_id(prefix: str, value: str) -> str:
    digest = hashlib.sha1(value.encode("utf-8")).hexdigest()[:8].upper()
    return f"{prefix}-{digest}"


def parse_scalar(raw: str) -> Any:
    raw = raw.strip()
    if not raw:
        return ""
    if raw.lower() == "true":
        return True
    if raw.lower() == "false":
        return False
    return raw.strip('"').strip("'")


def parse_frontmatter(text: str) -> tuple[dict[str, Any], str]:
    if not text.startswith("---\n"):
        return {}, text
    end = text.find("\n---", 4)
    if end < 0:
        return {}, text
    raw = text[4:end].splitlines()
    body = text[end + 4 :].strip()
    fm: dict[str, Any] = {}
    current_key: str | None = None
    for line in raw:
        if not line.strip():
            continue
        item = re.match(r"^\s*-\s+(.*)$", line)
        if item and current_key:
            fm.setdefault(current_key, [])
            if not isinstance(fm[current_key], list):
                fm[current_key] = [fm[current_key]]
            fm[current_key].append(parse_scalar(item.group(1)))
            continue
        pair = re.match(r"^([A-Za-z0-9_\-]+):\s*(.*)$", line)
        if pair:
            key, val = pair.groups()
            if val.strip():
                fm[key] = parse_scalar(val)
                current_key = None
            else:
                fm[key] = []
                current_key = key
    return fm, body


def load_spec() -> tuple[dict[str, list[dict[str, str]]], list[dict[str, str]]]:
    wb = load_workbook(SPEC_XLSX, data_only=True)
    fields_ws = wb["02_字段字典"]
    headers = [s(c.value) for c in fields_ws[1]]
    fields_by_table: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in fields_ws.iter_rows(min_row=2, values_only=True):
        rec = {headers[i]: s(row[i]) for i in range(min(len(headers), len(row)))}
        if not rec.get("业务表") or not rec.get("字段名称"):
            continue
        code = TABLE_CODE_BY_NAME.get(rec["业务表"], rec.get("业务表", ""))
        rec["表编号"] = code
        fields_by_table[code].append(rec)

    tables_ws = wb["01_表结构总表"]
    table_headers = [s(c.value) for c in tables_ws[1]]
    tables = []
    for row in tables_ws.iter_rows(min_row=2, values_only=True):
        rec = {table_headers[i]: s(row[i]) for i in range(min(len(table_headers), len(row)))}
        if rec.get("表编号", "").startswith("T"):
            tables.append(rec)
    return dict(fields_by_table), tables


def iter_notes() -> list[dict[str, Any]]:
    notes: list[dict[str, Any]] = []
    for path in sorted(VAULT_ROOT.rglob("*.md")):
        rel = path.relative_to(VAULT_ROOT)
        if any(part in EXCLUDE_PARTS or part.startswith(".") for part in rel.parts):
            continue
        text = path.read_text(encoding="utf-8", errors="ignore")
        fm, body = parse_frontmatter(text)
        title = s(fm.get("title")) or path.stem
        tags = fm.get("tags") or []
        if isinstance(tags, str):
            tags = re.split(r"[,，\s]+", tags)
        notes.append(
            {
                "path": str(rel),
                "parts": rel.parts,
                "frontmatter": fm,
                "body": body,
                "title": title,
                "tags": [s(t) for t in tags if s(t)],
                "type": s(fm.get("type")),
                "updated": s(fm.get("updated")) or s(fm.get("date")),
            }
        )
    return notes


def body_excerpt(note: dict[str, Any], limit: int = 900) -> str:
    return clean_text(note["body"], limit)


def lines_with_keywords(text: str, keywords: list[str], limit: int = 700) -> str:
    lines = []
    for line in text.splitlines():
        line_clean = clean_text(line, 220)
        if line_clean and any(k in line_clean for k in keywords):
            lines.append(line_clean)
        if len("；".join(lines)) >= limit:
            break
    return "；".join(dict.fromkeys(lines))[:limit]


def infer_series(title: str, fm: dict[str, Any]) -> str:
    source = f"{title} {s(fm.get('product_name'))} {s(fm.get('product_category'))}"
    for key in ["PK", "PC", "PL", "PY", "PJ"]:
        if key in source:
            return key
    if "咖啡" in source:
        return "PK 咖啡底胶烫金膜系列"
    if "塑胶" in source:
        return "PC 塑胶烫金膜系列"
    if "颜料" in source or "色箔" in source:
        return "PL-PY 颜料箔系列"
    return s(fm.get("product_category")) or "烫金膜/相关产品"


def infer_color(title: str) -> str:
    colors = ["亮金", "浅金", "古铜金", "金银", "银", "金", "彩色", "红", "蓝", "绿", "黑", "白", "哑面", "亮面"]
    return "、".join([c for c in colors if c in title][:3])


def note_base(note: dict[str, Any], status_note: str = "由脚本按规范表字段自动整理，需业务复核后发布") -> dict[str, str]:
    return {
        "来源Obsidian路径": note["path"],
        "原始标题": note["title"],
        "来源类型": note["type"],
        "标签": "、".join(note["tags"]),
        "更新时间": note["updated"],
        "入库备注": status_note,
    }


def is_quality_note(note: dict[str, Any]) -> bool:
    hay = f"{note['path']} {note['title']} {' '.join(note['tags'])} {s(note['frontmatter'].get('process_category'))} {s(note['frontmatter'].get('process_stage'))}"
    return bool(re.search(r"缺陷|问题|故障|异常|诊断|弊病|白点|缩孔|起皱|气泡|堵塞|附着力不足|掉粉|反粘|糊版|飞金", hay))


def category_for(note: dict[str, Any]) -> str:
    p = note["path"]
    t = note["type"]
    hay = f"{p} {t} {note['title']} {' '.join(note['tags'])}"
    if "10_报价数据库" in p or "quote" in t:
        return "T05_QUOTE"
    if "03_GitMemory_工艺配方客户案例知识库/产品条目" in p or "01_ProductSpec" in p or "product" in t:
        return "T01_PRODUCT"
    if "02_PRD_客户定制开发工单" in p or "L2_应用场景细分" in p:
        return "T02_SUBSTRATE"
    if "L3_底层工艺参数" in p or "11_涂布工艺知识资产" in p:
        return "T06_DEFECT" if is_quality_note(note) else "T04_PROCESS"
    if "04_L1L2L3_产品线应用工艺参数" in p:
        return "T03_COMPATIBILITY"
    if re.search(r"FAQ|话术|营销|提示词|Agent|智能体|Confluence", hay, re.I):
        return "T08_FAQ"
    if re.search(r"标签|tag|数据库视图", hay, re.I):
        return "T09_TAG"
    if re.search(r"供应链|供应商|采购", hay):
        return "T07_SUPPLY"
    return "T10_GOVERNANCE"


def map_note_to_row(note: dict[str, Any], code: str) -> dict[str, str]:
    fm = note["frontmatter"]
    title = note["title"]
    body = note["body"]
    excerpt = body_excerpt(note)
    base = note_base(note)

    if code == "T01_PRODUCT":
        product = s(fm.get("product_name")) or title
        return {
            **base,
            "产品编号": s(fm.get("product_id")) or stable_id("SKU", note["path"]),
            "产品系列": infer_series(title, fm),
            "颜色名称": s(fm.get("color_name")) or infer_color(title),
            "色号 / 内部代码": s(fm.get("color_code")) or re.sub(r"[^A-Za-z0-9\-]", "", product)[:30],
            "规格": s(fm.get("spec")) or lines_with_keywords(body, ["规格", "宽度", "长度", "厚度"], 260),
            "表面效果": s(fm.get("surface_effect")) or infer_color(f"{title} {excerpt}"),
            "产品状态": s(fm.get("status")) or "active",
            "产品图片": s(fm.get("product_image")) or s(fm.get("source_file")),
            "推荐应用行业": s(fm.get("application_scenarios")) or lines_with_keywords(body, ["应用", "场景", "行业", "包装", "礼盒"], 420),
            "推荐底材": s(fm.get("target_substrates")) or lines_with_keywords(body, ["适用底材", "纸", "PET", "PVC", "皮革", "膜"], 420),
            "不推荐底材": s(fm.get("not_recommended_substrates")) or lines_with_keywords(body, ["不推荐", "禁忌", "风险"], 320),
            "MOQ": s(fm.get("moq")),
            "常规交期": s(fm.get("lead_time")),
            "是否有库存": s(fm.get("has_stock")) or "需确认",
            "替代型号": s(fm.get("alternative_models")),
            "使用风险等级": s(fm.get("risk_level")) or ("中" if re.search(r"需.*核对|风险|禁忌", excerpt) else "需复核"),
            "必须打样场景": s(fm.get("must_test_scenarios")) or lines_with_keywords(body, ["打样", "测试", "复杂", "触感", "UV"], 360),
            "创建时间": note["updated"],
            "更新时间": note["updated"],
            "审核人": s(fm.get("source_quality")) or "数据运营初审",
        }
    if code == "T02_SUBSTRATE":
        substrate = s(fm.get("substrate_name")) or title
        return {
            **base,
            "底材名称": substrate,
            "底材分类": "、".join([x for x in ["纸张", "塑料", "皮革", "织物", "复合材料"] if x in f"{title}{excerpt}"]) or "需复核",
            "表面状态": lines_with_keywords(body, ["粗糙", "压纹", "平滑", "橘皮"], 220),
            "表面处理": s(fm.get("surface_treatment")) or lines_with_keywords(body, ["表面处理", "过胶", "UV", "触感", "电晕", "覆膜"], 360),
            "吸附难度": "高" if re.search(r"低表面能|粗糙|压纹|触感|风险", excerpt) else "中",
            "耐温情况": lines_with_keywords(body, ["温度", "℃", "耐温"], 220),
            "推荐膜系列": lines_with_keywords(body, ["推荐", "PK", "PC", "PL", "PY", "PJ"], 300),
            "高风险膜系列": lines_with_keywords(body, ["不推荐", "高风险", "谨慎"], 260),
            "常见应用": lines_with_keywords(body, ["应用", "行业", "场景", "包装", "礼盒"], 360),
            "常见问题": lines_with_keywords(body, ["问题", "风险", "掉粉", "反粘", "糊版", "飞金", "烫不实"], 500) or excerpt[:500],
            "处理建议": lines_with_keywords(body, ["建议", "补问", "预审", "测试", "打样", "确认"], 520) or "需结合客户实际材料打样确认",
            "审核状态": s(fm.get("status")) or "草稿",
        }
    if code == "T03_COMPATIBILITY":
        return {
            **base,
            "规则编号": stable_id("R", note["path"]),
            "产品型号": s(fm.get("product_name")) or infer_series(title, fm),
            "底材名称": s(fm.get("target_substrates")) or lines_with_keywords(body, ["底材", "纸", "PET", "PVC", "皮革"], 300),
            "表面处理": s(fm.get("surface_treatment")) or lines_with_keywords(body, ["UV", "触感", "覆膜", "电晕", "光油", "哑油"], 260),
            "适配等级": "推荐" if re.search(r"适合|推荐|通用|稳定", excerpt) else "视项目而定",
            "推荐理由": lines_with_keywords(body, ["适合", "推荐", "优势", "卖点", "稳定"], 520) or excerpt[:520],
            "风险说明": lines_with_keywords(body, ["风险", "禁忌", "需.*打样", "不推荐"], 420),
            "温度范围": lines_with_keywords(body, ["温度", "°C", "℃"], 260),
            "压力范围": lines_with_keywords(body, ["压力"], 180),
            "速度范围": lines_with_keywords(body, ["速度"], 180),
            "是否必须打样": "视项目而定",
            "关联实践云案例": title,
            "销售推荐话术": lines_with_keywords(body, ["对外表达", "话术", "客户"], 420),
            "审核人": "工艺/销售负责人复核",
        }
    if code == "T04_PROCESS":
        return {
            **base,
            "工艺名称": title,
            "适用产品": s(fm.get("product_name")) or "烫金膜涂布/热烫工艺",
            "温度范围": s(fm.get("temp_range")) or lines_with_keywords(body, ["温度", "℃", "°C"], 260),
            "压力范围": s(fm.get("pressure_range")) or lines_with_keywords(body, ["压力"], 220),
            "速度范围": s(fm.get("speed_range")) or lines_with_keywords(body, ["速度", "线速"], 220),
            "时间 / 停留时间": s(fm.get("dwell_time")) or lines_with_keywords(body, ["时间", "停留"], 160),
            "模具要求": lines_with_keywords(body, ["模头", "刮刀", "网穴", "唇口", "模具"], 420),
            "设备要求": lines_with_keywords(body, ["设备", "烘箱", "辊", "泵", "过滤", "纠偏"], 520),
            "环境要求": lines_with_keywords(body, ["环境", "洁净", "静电", "温湿度"], 360),
            "常见异常": s(fm.get("problem_tags")) or lines_with_keywords(body, ["异常", "缺陷", "问题", "故障", "弊病"], 520),
            "调机建议": lines_with_keywords(body, ["对策", "建议", "优化", "调整", "控制"], 650) or "见原文，需工艺负责人复核",
            "客户解释口径": lines_with_keywords(body, ["客户", "解释", "对外"], 360),
        }
    if code == "T05_QUOTE":
        prices = []
        for label, key in [("市场价RMB", "market_rmb"), ("市场价USD", "market_usd"), ("经销价RMB", "dealer_rmb"), ("经销价USD", "dealer_usd"), ("终端价RMB", "terminal_rmb")]:
            if s(fm.get(key)):
                prices.append(f"{label}:{s(fm.get(key))}")
        return {
            **base,
            "报价规则编号": s(fm.get("quote_id")) or stable_id("Q", note["path"]),
            "产品型号 / 系列": s(fm.get("product_name")) or title,
            "基础成本": "；".join(prices),
            "宽幅影响": s(fm.get("spec")),
            "数量阶梯": s(fm.get("customer_type")) or lines_with_keywords(body, ["数量", "阶梯", "客户类型"], 260),
            "损耗系数": "",
            "MOQ": s(fm.get("unit")),
            "交期规则": s(fm.get("effective_date")) or lines_with_keywords(body, ["交期", "生效"], 300),
            "加急费用": "",
            "定制费用": s(fm.get("package_info")),
            "价格等级": s(fm.get("currency")) or s(fm.get("customer_type")),
            "让步边界": "必须人工复核" if s(fm.get("human_review_required")).lower() == "true" else lines_with_keywords(body, ["人工", "复核", "边界"], 260),
            "替代方案": s(fm.get("quote_source")),
            "报价备注": lines_with_keywords(body, ["提醒", "规则", "使用说明", "报价"], 620) or excerpt[:620],
        }
    if code == "T06_DEFECT":
        causes = lines_with_keywords(body, ["原因", "可能原因", "形成", "导致"], 900)
        advice = lines_with_keywords(body, ["对策", "解决", "建议", "优化", "控制", "检查"], 900)
        return {
            **base,
            "问题编号": stable_id("D", note["path"]),
            "缺陷名称": title,
            "缺陷图片": s(fm.get("source_image")) or s(fm.get("renamed_file")),
            "可能原因 1": causes[:400] or excerpt[:400],
            "可能原因 2": causes[400:800],
            "可能原因 3": s(fm.get("problem_tags")),
            "调整建议": advice or "见原文诊断方案，需工艺/质检复核",
            "替代产品": "",
            "是否需重新打样": "视影响范围而定",
            "对客户解释": "该问题通常受材料、设备、工艺窗口和现场环境共同影响，建议先按来源知识卡完成排查并以实测结果确认。",
            "严重程度": "中" if "high" not in s(fm.get("hotfoil_relevance")) else "高",
            "审核状态": s(fm.get("status")) or "待审核",
        }
    if code == "T07_SUPPLY":
        return {
            **base,
            "供应商编号": stable_id("V", note["path"]),
            "供应商名称": s(fm.get("vendor_name")) or s(fm.get("company")) or title,
            "可供产品系列": s(fm.get("provided_products")) or lines_with_keywords(body, ["产品", "提供", "供应"], 420),
            "质量稳定等级": s(fm.get("quality_level")) or "需复核",
            "批次稳定性": lines_with_keywords(body, ["批次", "稳定"], 260),
            "常规交期": s(fm.get("lead_time")) or lines_with_keywords(body, ["交期", "供货"], 220),
            "最小采购量": s(fm.get("moq")),
            "供应风险": s(fm.get("supply_risk")) or "需复核",
            "替代供应商": "",
            "对外承诺限制": lines_with_keywords(body, ["承诺", "限制", "不能", "必须"], 420) or "需人工确认交期、库存与价格后对外承诺",
        }
    if code == "T08_FAQ":
        return {
            **base,
            "问题编号": stable_id("F", note["path"]),
            "客户常问问题": s(fm.get("question")) or title,
            "问题类型": "价格疑问" if "报价" in note["path"] else ("适配疑问" if re.search(r"底材|适配|工艺", excerpt) else "售前"),
            "中文回答": excerpt,
            "英文回答": "",
            "关联产品": s(fm.get("related_products")) or ("PINTE 品特烫金膜" if "烫金" in excerpt else ""),
            "关联实践案例": s(fm.get("related_cases")),
            "禁止承诺内容": s(fm.get("forbidden_promises")) or lines_with_keywords(body, ["禁止", "不能", "不得", "需人工复核"], 360),
            "适用客户阶段": s(fm.get("client_stage")) or "全流程",
        }
    if code == "T09_TAG":
        tag_name = title.replace("目录", "").strip() or stable_id("TAG", note["path"])
        return {
            **base,
            "标签编号": stable_id("TAG", note["path"]),
            "标签名称": tag_name,
            "标签类型": "工艺" if "工艺" in f"{title}{excerpt}" else ("产品系列" if re.search(r"PK|PC|PL|PY|PJ|产品", f"{title}{excerpt}") else "客户阶段"),
            "使用规则": excerpt,
            "上级标签": note["parts"][0] if note["parts"] else "",
            "同义词": s(fm.get("aliases")),
            "禁用词": "",
            "用于哪些应用": lines_with_keywords(body, ["应用", "推荐", "案例", "匹配"], 520),
        }
    return {
        **base,
        "知识编号": stable_id("GOV", note["path"]),
        "知识类型": note["parts"][0] if note["parts"] else "根目录",
        "知识标题": title,
        "知识正文": excerpt,
        "知识来源": note["path"],
        "来源可信度": s(fm.get("source_quality")) or ("高" if s(fm.get("status")) == "active" else "待验证"),
        "审核人": "数据运营初审",
        "审核状态": s(fm.get("status")) or "草稿",
        "版本号": s(fm.get("updated")) or "Obsidian同步",
        "更新时间": note["updated"],
        "失效条件": lines_with_keywords(body, ["失效", "过期", "更新"], 260),
        "使用次数": "",
        "反馈评分": "",
    }


def build_rows(fields_by_table: dict[str, list[dict[str, str]]], notes: list[dict[str, Any]]) -> dict[str, list[dict[str, str]]]:
    rows: dict[str, list[dict[str, str]]] = defaultdict(list)
    for note in notes:
        code = category_for(note)
        rows[code].append(map_note_to_row(note, code))

    # Add compatibility rows from product rows when the source explicitly contains substrate guidance.
    for note in notes:
        if "03_GitMemory_工艺配方客户案例知识库/产品条目" in note["path"] or "04_L1L2L3_产品线应用工艺参数" in note["path"]:
            rows["T03_COMPATIBILITY"].append(map_note_to_row(note, "T03_COMPATIBILITY"))

    # Add governance rows for every note so the full vault has an auditable AI-call registry.
    for note in notes:
        rows["T10_GOVERNANCE"].append(map_note_to_row(note, "T10_GOVERNANCE"))

    tag_sources: dict[str, dict[str, Any]] = {}
    for note in notes:
        for tag in note["tags"]:
            if not tag or tag in tag_sources:
                continue
            pseudo_note = {
                **note,
                "title": tag,
                "type": "frontmatter_tag",
                "body": f"标签来自 Obsidian frontmatter，用于知识云检索、路由和推荐。首个来源：{note['path']}",
                "tags": [tag],
            }
            tag_sources[tag] = pseudo_note
    for tag_note in sorted(tag_sources.values(), key=lambda n: n["title"]):
        rows["T09_TAG"].append(map_note_to_row(tag_note, "T09_TAG"))

    ordered: dict[str, list[dict[str, str]]] = {}
    for code in SHEET_TITLE_BY_CODE:
        field_names = [f["字段名称"] for f in fields_by_table.get(code, [])]
        filtered_rows = []
        for row in rows.get(code, []):
            rec = {col: s(row.get(col)) for col in field_names + META_COLUMNS}
            if any(rec.get(col) for col in field_names):
                filtered_rows.append(rec)
        ordered[code] = filtered_rows
    return ordered


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    fields_by_table, tables = load_spec()
    notes = iter_notes()
    rows_by_table = build_rows(fields_by_table, notes)
    summary = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "spec_xlsx": str(SPEC_XLSX),
        "vault_root": str(VAULT_ROOT),
        "note_count": len(notes),
        "top_directory_counts": Counter(n["parts"][0] if n["parts"] else "" for n in notes),
        "table_row_counts": {code: len(rows) for code, rows in rows_by_table.items()},
    }
    payload = {
        "fields_by_table": fields_by_table,
        "tables": tables,
        "sheet_title_by_code": SHEET_TITLE_BY_CODE,
        "meta_columns": META_COLUMNS,
        "rows_by_table": rows_by_table,
        "summary": summary,
    }
    (OUT_DIR / "knowledge_cloud_import_data.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
